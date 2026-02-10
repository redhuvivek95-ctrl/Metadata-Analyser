"""
Beewax Summon Data Extractor - Clean Version with Colored Participant Groups
Features:
- Multiple ID inputs with + button
- Extract all participant records
- Comparison table with color coding by search ID group
- Excel export with colors
- Fast mode (headless browser)
"""

import sys
import traceback

try:
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog
except ImportError as e:
    print(f"✗ Error importing Tkinter: {e}")
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    print("✗ Pandas not found. Install with: pip install pandas")
    sys.exit(1)

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.keys import Keys
except ImportError:
    print("✗ Selenium not found. Install with: pip install selenium")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    print("✗ openpyxl not found. Install with: pip install openpyxl")
    sys.exit(1)

import time
import threading
from datetime import datetime
import os
import re
import html
import subprocess
import platform


class LoadingDialog:
    """Loading dialog with spinner animation"""
    def __init__(self, parent):
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Loading...")
        self.dialog.geometry("350x150")
        self.dialog.configure(bg='#4A90E2')
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Center dialog
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (350 // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (150 // 2)
        self.dialog.geometry(f"350x150+{x}+{y}")
        
        self.spinner_label = tk.Label(self.dialog, text="⠋", 
                                     font=('Helvetica', 48), bg='#4A90E2', fg='white')
        self.spinner_label.pack(pady=20)
        
        self.status_label = tk.Label(self.dialog, text="Starting...",
                                     font=('Helvetica', 11), bg='#4A90E2', fg='white')
        self.status_label.pack()
        
        self.messages = [
            "🚀 Launching browser...",
            "🔍 Searching records...",
            "☕ Grab a coffee...",
            "✨ Working magic...",
            "⚡ Processing data...",
            "🎯 Almost there...",
        ]
        self.msg_index = 0
        self.animate()
    
    def animate(self):
        if self.dialog.winfo_exists():
            spinners = ['⠋', '⠙', '⠹', '⠸', '⠼', '⠴', '⠦', '⠧', '⠇', '⠏']
            current = self.spinner_label.cget('text')
            next_idx = (spinners.index(current) + 1) % len(spinners)
            self.spinner_label.config(text=spinners[next_idx])
            
            if next_idx == 0:
                self.msg_index = (self.msg_index + 1) % len(self.messages)
                self.status_label.config(text=self.messages[self.msg_index])
            
            self.dialog.after(100, self.animate)
    
    def update_status(self, message):
        if self.dialog.winfo_exists():
            self.status_label.config(text=message)
    
    def destroy(self):
        if self.dialog.winfo_exists():
            self.dialog.destroy()


class BeewaxExtractorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Beewax Summon Data Extractor Pro")
        
        # Set window size
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        self.root.geometry(f"{screen_width}x{screen_height}+0+0")
        
        # Configure style
        self.setup_styles()
        
        # Main Frame
        main_frame = tk.Frame(root, bg='#f5f7fa')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        # Initialize variables
        self.extracted_data = []
        self.driver = None
        self.is_running = False
        self.loading_dialog = None
        self.all_table_rows = []
        self.id_entries = []
        self.search_id_groups = {}  # Map search_id -> list of participant records
        
        # Color palette for different search ID groups
        self.group_colors = [
            '#FFE6E6',  # Light red
            '#E6F3FF',  # Light blue
            '#E6FFE6',  # Light green
            '#FFF0E6',  # Light orange
            '#F0E6FF',  # Light purple
            '#FFFFE6',  # Light yellow
            '#E6FFFF',  # Light cyan
            '#FFE6F0',  # Light pink
        ]
        
        # Build UI
        self.build_ui(main_frame)
        
        # Window close handler
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def setup_styles(self):
        """Setup ttk styles for modern appearance"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Modern Treeview styling
        style.configure("Custom.Treeview",
                       background="#ffffff",
                       foreground="#2C3E50",
                       rowheight=35,
                       fieldbackground="#ffffff",
                       borderwidth=0,
                       font=('Segoe UI', 9))
        
        style.configure("Custom.Treeview.Heading",
                       background="#3498DB",
                       foreground="white",
                       relief="flat",
                       borderwidth=0,
                       font=('Segoe UI', 10, 'bold'))
        
        style.map('Custom.Treeview',
                 background=[('selected', '#E8F4F8')],
                 foreground=[('selected', '#2C3E50')])
        
        style.map('Custom.Treeview.Heading',
                 background=[('active', '#2980B9')],
                 relief=[('active', 'flat')])
    
    def build_ui(self, parent):
        """Build the main UI with modern, attractive design"""
        
        # ==================== MODERN HEADER ====================
        header_frame = tk.Frame(parent, bg='#ffffff', relief=tk.FLAT)
        header_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Gradient-like header with shadow effect
        header_shadow = tk.Frame(header_frame, bg='#e0e0e0', height=4)
        header_shadow.pack(fill=tk.X)
        
        header_inner = tk.Frame(header_frame, bg='#2C3E50')
        header_inner.pack(fill=tk.X)
        
        # Title with modern styling - REDUCED WIDTH, NEW TITLE
        title_container = tk.Frame(header_inner, bg='#2C3E50')
        title_container.pack(pady=20)
        
        # Icon and title
        icon_label = tk.Label(title_container, text="🔬", font=('Segoe UI Emoji', 36), 
                             bg='#2C3E50')
        icon_label.pack(side=tk.LEFT, padx=10)
        
        title_text_frame = tk.Frame(title_container, bg='#2C3E50')
        title_text_frame.pack(side=tk.LEFT)
        
        tk.Label(title_text_frame, text="Data Extractor",
                 font=('Segoe UI', 22, 'bold'), bg='#2C3E50', fg='#ECF0F1').pack(anchor=tk.W)
        tk.Label(title_text_frame, text="Compare metadata across multiple records",
                 font=('Segoe UI', 10), bg='#2C3E50', fg='#BDC3C7').pack(anchor=tk.W, pady=(2, 0))
        
        # ==================== MAIN CONTENT AREA ====================
        content_frame = tk.Frame(parent, bg='#f5f7fa')
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # ==================== ID INPUT SECTION WITH INFO BOX ====================
        input_section = tk.Frame(content_frame, bg='#f5f7fa')
        input_section.pack(fill=tk.X, padx=15, pady=(0, 10))  # Reduced padding
        
        # Left side - ID inputs (70% width)
        left_panel = tk.Frame(input_section, bg='#ffffff', relief=tk.FLAT)
        left_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 15))
        
        # Card-like container with shadow
        id_card_shadow = tk.Frame(left_panel, bg='#e0e0e0', height=2)
        id_card_shadow.pack(fill=tk.X)
        
        id_card = tk.Frame(left_panel, bg='#ffffff')
        id_card.pack(fill=tk.BOTH, expand=True, padx=3, pady=3)
        
        # Section header
        id_header = tk.Frame(id_card, bg='#3498DB', height=50)
        id_header.pack(fill=tk.X)
        id_header.pack_propagate(False)
        
        tk.Label(id_header, text="🔍  Enter Record IDs", 
                font=('Segoe UI', 13, 'bold'), bg='#3498DB', fg='white').pack(side=tk.LEFT, padx=20, pady=12)
        
        # Scrollable ID container - REDUCED HEIGHT
        id_scroll_container = tk.Frame(id_card, bg='#ffffff')
        id_scroll_container.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)  # Reduced pady
        
        self.id_canvas = tk.Canvas(id_scroll_container, bg='#ffffff', height=180,  # Reduced from 250
                                   highlightthickness=0, bd=0)
        id_scrollbar = tk.Scrollbar(id_scroll_container, orient="vertical", 
                                    command=self.id_canvas.yview)
        self.id_container = tk.Frame(self.id_canvas, bg='#ffffff')
        
        self.id_container.bind("<Configure>",
            lambda e: self.id_canvas.configure(scrollregion=self.id_canvas.bbox("all")))
        
        self.id_canvas.create_window((0, 0), window=self.id_container, anchor="nw")
        self.id_canvas.configure(yscrollcommand=id_scrollbar.set)
        
        self.id_canvas.pack(side="left", fill="both", expand=True)
        id_scrollbar.pack(side="right", fill="y")
        
        # Add first 2 ID fields
        self.add_id_field()
        self.add_id_field()
        
        # Control buttons with modern design - REDUCED HEIGHT
        button_container = tk.Frame(id_card, bg='#ECF0F1', height=70)  # Reduced from 80
        button_container.pack(fill=tk.X)
        button_container.pack_propagate(False)
        
        button_inner = tk.Frame(button_container, bg='#ECF0F1')
        button_inner.pack(expand=True)
        
        # Add ID button
        add_id_btn = tk.Button(button_inner, text="➕  Add ID",
                              command=self.add_id_field,
                              bg='#27AE60', fg='white', font=('Segoe UI', 10, 'bold'),
                              relief=tk.FLAT, padx=20, pady=10, cursor='hand2',
                              activebackground='#229954', bd=0)
        add_id_btn.pack(side=tk.LEFT, padx=5)
        
        # Search button (larger, more prominent)
        self.search_btn = tk.Button(button_inner, text="🚀  Search & Extract",
                                    command=self.start_extraction,
                                    bg='#3498DB', fg='white', font=('Segoe UI', 12, 'bold'),
                                    relief=tk.FLAT, padx=30, pady=12, cursor='hand2',
                                    activebackground='#2980B9', bd=0)
        self.search_btn.pack(side=tk.LEFT, padx=10)
        
        # Export Excel button - WHITE TEXT
        self.export_xlsx_btn = tk.Button(button_inner, text="📊  Export to Excel",
                                         command=self.export_comparison_excel, state=tk.DISABLED,
                                         bg='#9B59B6', fg='white', font=('Segoe UI', 10, 'bold'),
                                         relief=tk.FLAT, padx=20, pady=10, cursor='hand2',
                                         activebackground='#8E44AD', bd=0)
        self.export_xlsx_btn.pack(side=tk.LEFT, padx=5)
        
        # Dummy export_btn for compatibility (removed from UI)
        self.export_btn = self.export_xlsx_btn  # Point to Excel button for any references
        
        # Clear All button
        self.clear_btn = tk.Button(button_inner, text="🗑️  Clear All",
                                   command=self.clear_results,
                                   bg='#E67E22', fg='white', font=('Segoe UI', 10, 'bold'),
                                   relief=tk.FLAT, padx=20, pady=10, cursor='hand2',
                                   activebackground='#D35400', bd=0)
        self.clear_btn.pack(side=tk.LEFT, padx=5)
        
        # Options frame - FAST MODE + EXCLUDE CATALOG/CITATION
        options_frame = tk.Frame(button_inner, bg='#ECF0F1')
        options_frame.pack(side=tk.LEFT, padx=15)
        
        self.headless_var = tk.BooleanVar(value=True)
        cb1 = tk.Checkbutton(options_frame, text="⚡ Fast Mode",
                            variable=self.headless_var, bg='#ECF0F1',
                            font=('Segoe UI', 9), activebackground='#ECF0F1',
                            selectcolor='#3498DB')
        cb1.pack(anchor=tk.W, pady=2)
        
        self.exclude_catalog_var = tk.BooleanVar(value=True)  # Default: exclude them
        cb2 = tk.Checkbutton(options_frame, text="🚫 Exclude Catalog/Citation",
                            variable=self.exclude_catalog_var, bg='#ECF0F1',
                            font=('Segoe UI', 9), activebackground='#ECF0F1',
                            selectcolor='#3498DB')
        cb2.pack(anchor=tk.W, pady=2)
        
        # Right side - Merge Rules Info Box (30% width)
        right_panel = tk.Frame(input_section, bg='#FFF9E6', relief=tk.FLAT, width=380)
        right_panel.pack(side=tk.RIGHT, fill=tk.Y)
        right_panel.pack_propagate(False)
        
        # Info card shadow
        info_shadow = tk.Frame(right_panel, bg='#e0e0e0', height=2)
        info_shadow.pack(fill=tk.X)
        
        info_card = tk.Frame(right_panel, bg='#FFF9E6')
        info_card.pack(fill=tk.BOTH, expand=True, padx=3, pady=3)
        
        # Info header with icon
        info_header = tk.Frame(info_card, bg='#F39C12', height=50)
        info_header.pack(fill=tk.X)
        info_header.pack_propagate(False)
        
        tk.Label(info_header, text="📚  Merge Rules Guide", 
                font=('Segoe UI', 12, 'bold'), bg='#F39C12', fg='white').pack(pady=12)
        
        # Scrollable info text
        info_scroll_frame = tk.Frame(info_card, bg='#FFF9E6')
        info_scroll_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        
        info_scroll = tk.Scrollbar(info_scroll_frame)
        info_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        info_text = tk.Text(info_scroll_frame, wrap=tk.WORD, font=('Segoe UI', 9),
                           bg='#FFF9E6', fg='#2C3E50', relief=tk.FLAT,
                           yscrollcommand=info_scroll.set, padx=10, pady=8,
                           height=12, bd=0)
        info_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        info_scroll.config(command=info_text.yview)
        
        # Add merge rules content with better formatting
        info_text.insert('1.0', "✅  Records MERGE if:\n", 'header_green')
        info_text.insert('end', "\n")
        info_text.insert('end', "📌 Same title + subtitle\n", 'bullet')
        info_text.insert('end', "📌 Same identifiers:\n", 'bullet')
        info_text.insert('end', "   • ISBN, EISBN (pub dates ±3 years)\n", 'sub_bullet')
        info_text.insert('end', "   • ISSN, EISSN\n", 'sub_bullet')
        info_text.insert('end', "   • SSID\n", 'sub_bullet')
        info_text.insert('end', "📌 Same publication year\n", 'bullet')
        info_text.insert('end', "📌 Same volume, issue, start page\n", 'bullet')
        info_text.insert('end', "\n\n")
        
        info_text.insert('end', "❌  Records will NOT merge if:\n", 'header_red')
        info_text.insert('end', "\n")
        info_text.insert('end', "🚫 Institutional repositories\n", 'bullet')
        info_text.insert('end', "🚫 Different languages\n", 'bullet')
        info_text.insert('end', "🚫 Different content types\n", 'bullet')
        info_text.insert('end', "🚫 Same-package, different URIs\n", 'bullet')
        info_text.insert('end', "🚫 Newspaper articles before 2000\n", 'bullet')
        
        info_text.tag_configure('header_green', font=('Segoe UI', 10, 'bold'), 
                               foreground='#27AE60', spacing1=5, spacing3=5)
        info_text.tag_configure('header_red', font=('Segoe UI', 10, 'bold'), 
                               foreground='#E74C3C', spacing1=5, spacing3=5)
        info_text.tag_configure('bullet', font=('Segoe UI', 9), 
                               foreground='#34495E', spacing1=3)
        info_text.tag_configure('sub_bullet', font=('Segoe UI', 8), 
                               foreground='#7F8C8D', spacing1=2)
        info_text.config(state=tk.DISABLED)
        
        # ==================== STATUS BAR ====================
        status_bar = tk.Frame(content_frame, bg='#34495E', height=45)  # Reduced from 50
        status_bar.pack(fill=tk.X, padx=15, pady=(0, 10))  # Reduced padding
        status_bar.pack_propagate(False)
        
        status_inner = tk.Frame(status_bar, bg='#34495E')
        status_inner.pack(fill=tk.X, padx=20, pady=10)  # Reduced pady
        
        self.status_label = tk.Label(status_inner, text="📊 Ready to extract data",
                                     font=('Segoe UI', 10), bg='#34495E', fg='#ECF0F1', 
                                     anchor=tk.W)
        self.status_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        self.record_count_label = tk.Label(status_inner, text="Records: 0",
                                          font=('Segoe UI', 10, 'bold'), bg='#34495E',
                                          fg='#3498DB', anchor=tk.E)
        self.record_count_label.pack(side=tk.RIGHT, padx=15)
        
        # ==================== RESULTS SECTION ====================
        results_section = tk.Frame(content_frame, bg='#ffffff')
        results_section.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 10))  # Reduced padding
        
        # Results header
        results_header_frame = tk.Frame(results_section, bg='#3498DB', height=50)
        results_header_frame.pack(fill=tk.X)
        results_header_frame.pack_propagate(False)
        
        results_header_left = tk.Frame(results_header_frame, bg='#3498DB')
        results_header_left.pack(side=tk.LEFT, fill=tk.Y, padx=20, pady=10)
        
        tk.Label(results_header_left, text="📋 Extraction Results",
                 font=('Segoe UI', 13, 'bold'), bg='#3498DB', fg='white').pack(side=tk.LEFT)
        
        tk.Label(results_header_left, text=" • Color-coded by record group",
                 font=('Segoe UI', 9), bg='#3498DB', fg='#ECF0F1').pack(side=tk.LEFT, padx=10)
        
        # Filter controls
        filter_frame = tk.Frame(results_header_frame, bg='#3498DB')
        filter_frame.pack(side=tk.RIGHT, padx=20, pady=10)
        
        tk.Label(filter_frame, text="Filter:", font=('Segoe UI', 9, 'bold'),
                bg='#3498DB', fg='white').pack(side=tk.LEFT, padx=(0, 8))
        
        self.filter_var = tk.StringVar(value="All")
        self.filter_dropdown = ttk.Combobox(filter_frame, textvariable=self.filter_var,
                                           values=["All", "Match", "Mismatch", "No data"],
                                           state="readonly", width=12, font=('Segoe UI', 9))
        self.filter_dropdown.pack(side=tk.LEFT)
        self.filter_dropdown.bind('<<ComboboxSelected>>', self.apply_filter)
        
        # Table container - MAXIMUM SIZE (no buttons or log below)
        table_container = tk.Frame(results_section, bg='#ecf0f1', height=600)  # Increased from 350 to 600!
        table_container.pack(fill=tk.BOTH, expand=True, padx=3, pady=3)
        table_container.pack_propagate(False)
        
        # Table with scrollbars
        table_inner = tk.Frame(table_container, bg='#ffffff')
        table_inner.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        self.table = ttk.Treeview(table_inner, show="headings", style="Custom.Treeview")
        self.table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar_y = ttk.Scrollbar(table_inner, orient=tk.VERTICAL, command=self.table.yview)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.table.configure(yscrollcommand=scrollbar_y.set)
        
        scrollbar_x = ttk.Scrollbar(results_section, orient=tk.HORIZONTAL, command=self.table.xview)
        scrollbar_x.pack(fill=tk.X, pady=(3, 0))
        self.table.configure(xscrollcommand=scrollbar_x.set)
        
        # NO ACTION BUTTONS OR ACTIVITY LOG BELOW - ALL MOVED TO TOP
        
        self.log("🎉 Application started successfully!")
        self.log("✨ Enter Record IDs and click 'Search & Extract' to begin")
    
    def add_id_field(self):
        """Add a new ID entry field with modern styling"""
        id_frame = tk.Frame(self.id_container, bg='#ffffff', relief=tk.SOLID, bd=1)
        id_frame.pack(fill=tk.X, pady=5, padx=5)  # Increased pady from 4 to 5
        
        num = len(self.id_entries) + 1
        
        # Get color for this record
        color_idx = (num - 1) % len(self.group_colors)
        record_color = self.group_colors[color_idx]
        
        # Color indicator bar
        color_bar = tk.Frame(id_frame, bg=record_color, width=8)
        color_bar.pack(side=tk.LEFT, fill=tk.Y)
        
        # Label with modern styling
        label_frame = tk.Frame(id_frame, bg='#ECF0F1', width=80)
        label_frame.pack(side=tk.LEFT, fill=tk.Y)
        label_frame.pack_propagate(False)
        
        tk.Label(label_frame, text=f"Record {num}", bg='#ECF0F1', fg='#2C3E50',
                font=('Segoe UI', 9, 'bold')).pack(expand=True)
        
        # Entry box with INCREASED HEIGHT
        entry_container = tk.Frame(id_frame, bg='#ffffff')
        entry_container.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10, pady=10)  # Increased pady
        
        entry = tk.Entry(entry_container, font=('Segoe UI', 11), relief=tk.FLAT,  # Increased font size
                        bd=0, bg='#F8F9FA', fg='#2C3E50',
                        insertbackground='#3498DB')
        entry.pack(fill=tk.X, ipady=8)  # Increased ipady from 6 to 8
        entry.bind('<Return>', lambda e: self.start_extraction())
        
        # Hover effect
        def on_enter(e):
            entry.config(bg='#E8F4F8')
        def on_leave(e):
            entry.config(bg='#F8F9FA')
        entry.bind('<Enter>', on_enter)
        entry.bind('<Leave>', on_leave)
        
        if len(self.id_entries) >= 1:
            remove_btn = tk.Button(id_frame, text="✕",
                                  command=lambda f=id_frame, e=entry: self.remove_id_field(f, e),
                                  bg='#E74C3C', fg='white', relief=tk.FLAT,
                                  font=('Segoe UI', 10, 'bold'), cursor='hand2',  # Increased font
                                  width=4, height=1, bd=0,
                                  activebackground='#C0392B')
            remove_btn.pack(side=tk.RIGHT, padx=10)  # Increased padx
        
        self.id_entries.append(entry)
        self.id_container.update_idletasks()
        self.id_canvas.configure(scrollregion=self.id_canvas.bbox("all"))
    
    def remove_id_field(self, frame, entry):
        """Remove an ID entry field"""
        if len(self.id_entries) > 1:
            self.id_entries.remove(entry)
            frame.destroy()
            
            # Renumber and recolor
            for i, frame in enumerate(self.id_container.winfo_children(), 1):
                label_frame = frame.winfo_children()[0]
                label = label_frame.winfo_children()[0]
                label.config(text=f"ID {i}:")
                
                # Update color
                color_idx = (i - 1) % len(self.group_colors)
                record_color = self.group_colors[color_idx]
                label_frame.config(bg=record_color)
                label.config(bg=record_color)
            
            self.id_container.update_idletasks()
            self.id_canvas.configure(scrollregion=self.id_canvas.bbox("all"))
    
    def log(self, message):
        """Log a message - now just prints to console"""
        timestamp = datetime.now().strftime('%H:%M:%S')
        log_msg = f"[{timestamp}] {message}"
        print(log_msg)
    
    def update_status(self, message):
        self.root.after(0, lambda: self.status_label.config(text=f"📊 Status: {message}"))
    
    def show_loading(self):
        """Show loading dialog"""
        self.loading_dialog = LoadingDialog(self.root)
    
    def hide_loading(self):
        """Hide loading dialog"""
        if self.loading_dialog:
            self.loading_dialog.destroy()
            self.loading_dialog = None
    
    def update_loading_status(self, message):
        """Update loading dialog status"""
        if self.loading_dialog:
            self.loading_dialog.update_status(message)
    
    def start_extraction(self):
        """Start extraction for ALL entered IDs"""
        if self.is_running:
            messagebox.showwarning("Already Running", "Extraction is already in progress.")
            return
        
        search_ids = [entry.get().strip() for entry in self.id_entries if entry.get().strip()]
        
        if not search_ids:
            messagebox.showwarning("Input Required", "Please enter at least one ID.")
            return
        
        self.is_running = True
        self.search_btn.config(state=tk.DISABLED, bg='#cccccc')
        self.export_btn.config(state=tk.DISABLED, bg='#cccccc')
        self.export_xlsx_btn.config(state=tk.DISABLED, bg='#cccccc')
        self.log(f"Starting extraction for {len(search_ids)} ID(s)...")
        
        self.extracted_data = []
        self.search_id_groups = {}
        self.clear_table()
        
        self.root.after(100, self.show_loading)
        
        thread = threading.Thread(target=self.extract_data_multiple, args=(search_ids,), daemon=True)
        thread.start()
    
    def extract_data_multiple(self, search_ids):
        """Extract data for multiple IDs"""
        try:
            all_participants = []
            
            for idx, search_id in enumerate(search_ids, 1):
                self.log(f"🔍 Processing ID {idx}/{len(search_ids)}: {search_id}")
                self.update_status(f"Processing ID {idx}/{len(search_ids)}")
                self.update_loading_status(f"Extracting ID {idx} of {len(search_ids)}...")
                
                participants = self.extract_single_id(search_id)
                
                if participants:
                    self.log(f"✅ Found {len(participants)} record(s) for {search_id}")
                    all_participants.extend(participants)
                    # Group participants by search ID
                    self.search_id_groups[search_id] = participants
                else:
                    self.log(f"⚠️ No records found for {search_id}")
                    self.search_id_groups[search_id] = []
            
            if not all_participants:
                self.root.after(0, self.hide_loading)
                self.update_status("No data extracted")
                self.root.after(0, lambda: messagebox.showwarning(
                    "No Data",
                    f"Could not extract any records from {len(search_ids)} ID(s).\n\n"
                    "Please check:\n"
                    "1. IDs are correct\n"
                    "2. You have access to Beewax\n"
                    "3. Records exist for these IDs"
                ))
                return
            
            self.extracted_data = all_participants
            self.root.after(0, lambda: self.record_count_label.config(
                text=f"Records: {len(all_participants)}"))
            
            self.root.after(0, self.hide_loading)
            self.root.after(0, self.populate_comparison_table)
            
            self.update_status(f"✅ Complete! Extracted {len(all_participants)} records")
            self.root.after(0, lambda: self.export_btn.config(state=tk.NORMAL, bg='#9B59B6'))
            self.root.after(0, lambda: self.export_xlsx_btn.config(state=tk.NORMAL, bg='#9B59B6'))
            self.log(f"🎉 Extraction complete! {len(all_participants)} records from {len(search_ids)} ID(s)")
            
        except Exception as e:
            self.root.after(0, self.hide_loading)
            self.log(f"❌ ERROR: {e}")
            self.log(traceback.format_exc())
            self.update_status("Error occurred")
            self.root.after(0, lambda: messagebox.showerror("Error",
                                                            f"An error occurred:\n\n{str(e)}\n\nCheck the log for details."))
        finally:
            self.is_running = False
            self.root.after(0, lambda: self.search_btn.config(state=tk.NORMAL, bg='#2196F3'))
    
    def extract_single_id(self, search_id):
        """Extract all participant records for a single ID"""
        try:
            options = webdriver.ChromeOptions()
            if self.headless_var.get():
                options.add_argument('--headless=new')
                options.add_argument('--disable-gpu')
            
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-images')
            options.page_load_strategy = 'eager'
            
            driver = webdriver.Chrome(options=options)
            driver.implicitly_wait(5)
            
            self.log(f"Navigating to Beewax for {search_id}")
            driver.get("https://beewax.summon.serialssolutions.com/apps/stui/SearchResults")
            time.sleep(1.5)
            
            wait = WebDriverWait(driver, 15)
            
            # Find search box
            search_box = None
            selectors = [
                ("CSS", "input[name='q']"),
                ("CSS", "input[type='search']"),
                ("CSS", "input[type='text']"),
            ]
            
            for selector_type, selector in selectors:
                try:
                    if selector_type == "CSS":
                        search_box = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
                    self.log(f"Found search box using {selector}")
                    break
                except:
                    continue
            
            if not search_box:
                self.log(f"Could not find search box for {search_id}")
                driver.quit()
                return []
            
            search_box.clear()
            search_box.send_keys(search_id)
            search_box.send_keys(Keys.RETURN)
            time.sleep(2)
            
            # Find participant links
            all_links = driver.find_elements(By.TAG_NAME, 'a')
            participant_data = []
            excluded_count = 0
            
            for link in all_links:
                try:
                    href = link.get_attribute('href')
                    if not href:
                        continue
                    
                    record_id = None
                    if 'sys_participant_id=' in href:
                        record_id = href.split('sys_participant_id=')[1].split('&')[0]
                    elif 'GetPhysicalRecord?id=' in href:
                        record_id = href.split('GetPhysicalRecord?id=')[1].split('&')[0]
                    else:
                        continue
                    
                    # Filter out unwanted records - ONLY IF CHECKBOX IS ENABLED
                    if self.exclude_catalog_var.get():  # Check if exclusion is enabled
                        record_lower = record_id.lower()
                        if any(pattern in record_lower for pattern in [
                            '_txt', 'citationcount', 'citationtrail', '_catalog', '_citation'
                        ]):
                            excluded_count += 1
                            continue
                    
                    if (record_id, href) not in participant_data:
                        participant_data.append((record_id, href))
                except:
                    continue
            
            if excluded_count > 0 and self.exclude_catalog_var.get():
                self.log(f"Excluded {excluded_count} records (citation/catalog/txt)")
            
            self.log(f"Found {len(participant_data)} valid participant records")
            
            if not participant_data:
                driver.quit()
                return []
            
            # Extract metadata from each participant
            records = []
            for idx, (record_id, url) in enumerate(participant_data, 1):
                self.log(f"Extracting {idx}/{len(participant_data)}: {record_id[:40]}...")
                driver.get(url)
                time.sleep(1.5)
                
                data = self.extract_details(record_id, driver.page_source)
                data['Record_ID'] = record_id
                data['Search_ID'] = search_id  # Track which search this came from
                records.append(data)
            
            driver.quit()
            return records
            
        except Exception as e:
            self.log(f"Error extracting {search_id}: {e}")
            if 'driver' in locals():
                try:
                    driver.quit()
                except:
                    pass
            return []
    
    def extract_details(self, record_id, page_source):
        """Extract metadata from page source"""
        data = {}
        data['Record_ID'] = record_id
        
        field_mappings = {
            'Title': ['Title', 'Title_t'],
            'Subtitle': ['Subtitle', 'Subtitle_t'],
            'Author': ['Author', 'Author_t'],
            'PublicationTitle': ['PublicationTitle', 'PublicationTitle_t'],
            'DOI': ['DOI', 'DOI_t'],
            'ISBN': ['ISBN', 'ISBN_t'],
            'EISBN': ['EISBN', 'EISBN_t'],
            'ISSN': ['ISSN', 'ISSN_t'],
            'EISSN': ['EISSN', 'EISSN_t'],
            'Volume': ['Volume', 'Volume_t'],
            'Issue': ['Issue', 'Issue_t'],
            'StartPage': ['StartPage', 'StartPage_t'],
            'EndPage': ['EndPage', 'EndPage_t'],
            'PublicationDateText': ['PublicationDateText', 'PublicationDate_t'],
            'PublicationDateYear': ['PublicationDateYear_s', 'PublicationYear'],
            'IsOpenAccess': ['IsOpenAccess', 'IsOpenAccess_b'],
            'LanguageEffective': ['LanguageEffective', 'LanguageEffective_s'],
            'Language_s': ['Language_s', 'Language'],
            'ContentType': ['ContentType', 'ContentType_t', 'ContentType_s110'],
            'SSID': ['SSID', 'SSID_t'],
            'URI': ['URI', 'URI_t', 'URI_s'],
        }
        
        ps = html.unescape(page_source)
        
        for display_field, search_fields in field_mappings.items():
            value = ""
            
            for field_variant in search_fields:
                pattern = rf'\b{re.escape(field_variant)}\s*[:\t]\s*([^\n\r<]{{1,200}}?)(?:\n|<|$)'
                matches = re.finditer(pattern, ps, flags=re.IGNORECASE | re.MULTILINE)
                for m in matches:
                    val = m.group(1).strip()
                    val = re.sub(r'<[^>]+>', '', val).strip()
                    val = val.split('\t')[0].strip()
                    if val and len(val) > 0:
                        value = val
                        break
                
                if value:
                    break
            
            data[display_field] = value
        
        return data
    
    def populate_comparison_table(self):
        """Populate table with color-coded groups"""
        self.clear_table()
        
        if not self.extracted_data:
            return
        
        # Get ordered list of search IDs (from ID entry order)
        ordered_search_ids = [entry.get().strip() for entry in self.id_entries if entry.get().strip()]
        
        # Build columns with color-coded headers
        participant_ids = []
        participant_search_mapping = {}  # Map participant_id -> search_id
        
        for search_id in ordered_search_ids:
            if search_id in self.search_id_groups:
                for rec in self.search_id_groups[search_id]:
                    pid = rec['Record_ID']
                    participant_ids.append(pid)
                    participant_search_mapping[pid] = search_id
        
        fields = sorted({k for r in self.extracted_data for k in r.keys() if k not in ['Record_ID', 'Search_ID']})
        
        columns = ["Field"] + participant_ids + ["Match"]
        self.table["columns"] = columns
        
        self.table.heading("Field", text="Field")
        self.table.column("Field", width=200, anchor=tk.W, stretch=False)
        
        for pid in participant_ids:
            display_name = pid[:30] + "..." if len(pid) > 30 else pid
            self.table.heading(pid, text=display_name)
            self.table.column(pid, width=350, anchor=tk.W, stretch=True)
        
        self.table.heading("Match", text="Match")
        self.table.column("Match", width=120, anchor=tk.CENTER, stretch=False)
        
        self.all_table_rows = []
        
        # Add header row showing which group each participant belongs to
        header_values = ["Record Group →"]
        search_id_index = {sid: idx for idx, sid in enumerate(ordered_search_ids)}
        
        for pid in participant_ids:
            search_id = participant_search_mapping.get(pid, "Unknown")
            record_num = search_id_index.get(search_id, 0) + 1
            header_values.append(f"Record {record_num}")
        header_values.append("")
        
        header_row = {
            'type': 'header',
            'values': header_values,
            'tags': ("header",),
            'colors': {}
        }
        
        # Map colors to columns
        for col_idx, pid in enumerate(participant_ids, 1):
            search_id = participant_search_mapping.get(pid)
            if search_id:
                group_idx = search_id_index.get(search_id, 0)
                color = self.group_colors[group_idx % len(self.group_colors)]
                header_row['colors'][col_idx] = color
        
        self.all_table_rows.append(header_row)
        self.table.insert("", tk.END, values=header_values, tags=("header",))
        
        # Add data rows with matching color per participant group
        for field in fields:
            # Create value mapping
            value_map = {}
            for rec in self.extracted_data:
                pid = rec.get('Record_ID')
                value_map[pid] = (rec.get(field) or "").strip()
            
            values = [value_map.get(pid, "") for pid in participant_ids]
            
            # Special handling for URI field - no match/mismatch comparison
            if field == 'URI':
                match_mark = ""  # No comparison for URI
                tags = ("uri_field",)  # Special tag for URI
            else:
                all_empty = all(not v or not str(v).strip() for v in values)
                
                if all_empty:
                    match_mark = "ℹ️ No data"
                    tags = ("nodata",)
                else:
                    match = self.values_match(values)
                    match_mark = "✅ Match" if match else "❌ Mismatch"
                    tags = ("match",) if match else ("mismatch",)
            
            row_values = [field] + values + [match_mark]
            
            row_data = {
                'type': 'data',
                'values': row_values,
                'tags': tags,
                'colors': {}
            }
            
            # Apply group colors
            for col_idx, pid in enumerate(participant_ids, 1):
                search_id = participant_search_mapping.get(pid)
                if search_id:
                    group_idx = search_id_index.get(search_id, 0)
                    color = self.group_colors[group_idx % len(self.group_colors)]
                    row_data['colors'][col_idx] = color
            
            self.all_table_rows.append(row_data)
            
            # Note: Tkinter Treeview doesn't support per-cell colors easily
            # We'll handle this in Excel export, but show basic tags in UI
            self.table.insert("", tk.END, values=row_values, tags=tags)
        
        self.table.tag_configure("header", background="#4A90E2", foreground="white", 
                                font=('Helvetica', 11, 'bold'))
        self.table.tag_configure("match", background="#C8E6C9", foreground="#1B5E20")
        self.table.tag_configure("mismatch", background="#FFE082", foreground="#E65100")
        self.table.tag_configure("nodata", background="#BBDEFB", foreground="#1565C0")
        self.table.tag_configure("uri_field", background="#FFFFFF", foreground="#2C3E50")  # URI: neutral white background
    
    def values_match(self, values):
        """Check if values match"""
        def normalize(v):
            if not v:
                return ""
            v = str(v).strip().lower()
            v = ' '.join(v.split())
            return v
        
        non_empty = [normalize(v) for v in values if v and str(v).strip()]
        
        if not non_empty:
            return True
        if len(non_empty) == 1:
            return True
        
        first = non_empty[0]
        return all(v == first for v in non_empty)
    
    def apply_filter(self, event=None):
        """Filter table rows"""
        filter_value = self.filter_var.get()
        
        for item in self.table.get_children():
            self.table.delete(item)
        
        for row_data in self.all_table_rows:
            row_type = row_data['type']
            values = row_data['values']
            tags = row_data['tags']
            
            if row_type == 'header':
                self.table.insert("", tk.END, values=values, tags=tags)
                continue
            
            # Special handling for URI field
            if 'uri_field' in tags:
                # Get the field name (first value in row)
                field_name = values[0] if values else ""
                
                # Check if all URI values are empty
                uri_values = values[1:-1]  # All values except field name and match column
                all_empty = all(not v or not str(v).strip() for v in uri_values)
                
                # Show URI only in "All" filter or "No data" filter if URIs are empty
                if filter_value == "All":
                    self.table.insert("", tk.END, values=values, tags=tags)
                elif filter_value == "No data" and all_empty:
                    self.table.insert("", tk.END, values=values, tags=tags)
                # Skip URI for "Match" and "Mismatch" filters
                continue
            
            # Regular field filtering
            if filter_value == "All":
                self.table.insert("", tk.END, values=values, tags=tags)
            elif filter_value == "Match" and 'match' in tags:
                self.table.insert("", tk.END, values=values, tags=tags)
            elif filter_value == "Mismatch" and 'mismatch' in tags:
                self.table.insert("", tk.END, values=values, tags=tags)
            elif filter_value == "No data" and 'nodata' in tags:
                self.table.insert("", tk.END, values=values, tags=tags)
    
    def clear_table(self):
        """Clear table"""
        for item in self.table.get_children():
            self.table.delete(item)
        self.table["columns"] = []
        self.all_table_rows = []
    
    def clear_ids(self):
        """Clear all ID entries"""
        for entry in self.id_entries:
            entry.delete(0, tk.END)
        self.log("🗑️ ID fields cleared")
    
    def clear_results(self):
        """Clear everything - table, data, and ID fields"""
        self.clear_table()
        self.extracted_data = []
        self.search_id_groups = {}
        
        # Clear all ID entry fields
        for entry in self.id_entries:
            entry.delete(0, tk.END)
        
        self.filter_var.set("All")
        self.export_btn.config(state=tk.DISABLED, bg='#9B59B6')  # Restore purple color
        self.export_xlsx_btn.config(state=tk.DISABLED, bg='#9B59B6')
        self.record_count_label.config(text="Records: 0")
        self.update_status("Ready")
        self.log("🗑️ All cleared. Ready for new search.")
    
    def export_to_csv(self):
        """Export to CSV"""
        if not self.extracted_data:
            messagebox.showwarning("No Data", "No data to export.")
            return
        
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialfile=f"beewax_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            if not filename:
                return
            
            df = pd.DataFrame(self.extracted_data)
            df.to_csv(filename, index=False)
            
            self.log(f"✅ CSV exported to: {filename}")
            messagebox.showinfo("Success", f"CSV exported successfully!\n\n{filename}")
        except Exception as e:
            self.log(f"❌ CSV export error: {e}")
            messagebox.showerror("Export Error", f"Failed to export CSV:\n{str(e)}")
    
    def export_comparison_excel(self):
        """Export comparison to Excel with color-coded groups"""
        if not self.extracted_data:
            messagebox.showwarning("No Data", "No data to export.")
            return
        
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"beewax_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            if not filename:
                return
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Comparison"
            
            # Color fills
            match_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            mismatch_fill = PatternFill(start_color="FFE082", end_color="FFE082", fill_type="solid")
            nodata_fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            
            # Group colors for participants
            ordered_search_ids = [entry.get().strip() for entry in self.id_entries if entry.get().strip()]
            search_id_index = {sid: idx for idx, sid in enumerate(ordered_search_ids)}
            
            participant_ids = []
            participant_search_mapping = {}
            
            for search_id in ordered_search_ids:
                if search_id in self.search_id_groups:
                    for rec in self.search_id_groups[search_id]:
                        pid = rec['Record_ID']
                        participant_ids.append(pid)
                        participant_search_mapping[pid] = search_id
            
            fields = sorted({k for r in self.extracted_data for k in r.keys() if k not in ['Record_ID', 'Search_ID']})
            
            # Write header row with participant IDs and colors
            ws.cell(1, 1, "Field").font = header_font
            ws.cell(1, 1).fill = header_fill
            
            for col, pid in enumerate(participant_ids, 2):
                cell = ws.cell(1, col, pid[:30])
                cell.font = header_font
                
                # Apply group color to header
                search_id = participant_search_mapping.get(pid)
                if search_id:
                    group_idx = search_id_index.get(search_id, 0)
                    color_hex = self.group_colors[group_idx % len(self.group_colors)].replace('#', '')
                    cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                else:
                    cell.fill = header_fill
                
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 35
            
            match_col = len(participant_ids) + 2
            ws.cell(1, match_col, "Match").font = header_font
            ws.cell(1, match_col).fill = header_fill
            ws.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 25
            
            # Write data rows
            for row, field in enumerate(fields, 2):
                ws.cell(row, 1, field)
                
                value_map = {}
                for rec in self.extracted_data:
                    pid = rec.get('Record_ID')
                    value_map[pid] = (rec.get(field) or "").strip()
                
                values = [value_map.get(pid, "") for pid in participant_ids]
                
                for col, (val, pid) in enumerate(zip(values, participant_ids), 2):
                    cell = ws.cell(row, col, val)
                    
                    # Apply group background color
                    search_id = participant_search_mapping.get(pid)
                    if search_id:
                        group_idx = search_id_index.get(search_id, 0)
                        color_hex = self.group_colors[group_idx % len(self.group_colors)].replace('#', '')
                        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                
                # Special handling for URI field - no match/mismatch
                if field == 'URI':
                    ws.cell(row, match_col, "")  # Empty match column for URI
                else:
                    all_empty = all(not v or not str(v).strip() for v in values)
                    
                    if all_empty:
                        ws.cell(row, match_col, "ℹ️ No data")
                        ws.cell(row, match_col).fill = nodata_fill
                    else:
                        match = self.values_match(values)
                        ws.cell(row, match_col, "✅" if match else "❌")
                        ws.cell(row, match_col).fill = match_fill if match else mismatch_fill
            
            ws.freeze_panes = ws['B2']
            wb.save(filename)
            
            self.log(f"✅ Excel exported to: {filename}")
            
            try:
                if platform.system() == 'Windows':
                    os.startfile(filename)
                elif platform.system() == 'Darwin':
                    subprocess.call(['open', filename])
                else:
                    subprocess.call(['xdg-open', filename])
                messagebox.showinfo("Success", f"Excel exported and opened!\n\n{filename}")
            except:
                messagebox.showinfo("Success", f"Excel exported!\n\n{filename}")
        
        except Exception as e:
            self.log(f"❌ Excel export error: {e}")
            messagebox.showerror("Export Error", f"Failed to export Excel:\n{str(e)}")
    
    def on_closing(self):
        """Handle window close"""
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
        self.root.destroy()


def main():
    print("\n" + "=" * 70)
    print("  🔬 Beewax Summon Data Extractor Pro")
    print("=" * 70 + "\n")
    
    try:
        root = tk.Tk()
        app = BeewaxExtractorGUI(root)
        root.mainloop()
    except Exception as e:
        print(f"\n❌ FATAL ERROR: {e}\n")
        traceback.print_exc()
        input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()
